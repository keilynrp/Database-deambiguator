"""
Enterprise Excel exporter — multi-sheet branded workbook.
Used by POST /exports/excel in backend/routers/reports.py.
"""
from __future__ import annotations

import logging
from io import BytesIO
from typing import List

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from backend import models
from backend.tenant_access import scope_query_to_org

logger = logging.getLogger(__name__)

# Violet brand palette (hex, no #)
_HEADER_FG   = "5B21B6"   # violet-800
_HEADER_FONT = "FFFFFF"   # white

_HEADER_FILL = PatternFill("solid", fgColor=_HEADER_FG)
_HEADER_FONT_STYLE = Font(color=_HEADER_FONT, bold=True, size=11)
_SUBHEADER_FONT = Font(bold=True, size=10)
_WRAP_TOP = Alignment(wrap_text=True, vertical="top")
_CAVEAT_FONT = Font(italic=True, size=9, color="FF6B7280")


def _style_header_row(ws, cols: list[str], row: int = 1) -> None:
    """Write and style a header row with violet fill + white bold text.

    `row` exists so a sheet can put a caveat line above its header (5.3) and
    still freeze the header rather than the caveat.
    """
    for col_idx, header in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT_STYLE
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = f"A{row + 1}"


def _autofit(ws, min_width: int = 10, max_width: int = 50) -> None:
    """Approximate column width based on content length."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                max_len = max(max_len, cell_len)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 2, max_width))


class EnterpriseExcelExporter:
    """Build a branded multi-sheet Excel workbook and return raw bytes."""

    _ENTITY_CAP = 5_000

    def _entities_query(self, db: Session, domain_id: str, org_id: int | None):
        query = scope_query_to_org(db.query(models.RawEntity), models.RawEntity, org_id)
        if domain_id:
            query = query.filter(models.RawEntity.domain == domain_id)
        return query

    def _harmonization_query(self, db: Session, org_id: int | None):
        return scope_query_to_org(db.query(models.HarmonizationLog), models.HarmonizationLog, org_id)

    def build(
        self,
        db: Session,
        domain_id: str,
        sections: List[str],
        org_id: int | None = None,
        manual_sections: list[dict[str, str]] | None = None,
    ) -> bytes:
        wb = openpyxl.Workbook()

        # ── Sheet 1: Summary KPIs ──────────────────────────────────────────────
        ws_summary = wb.active
        assert ws_summary is not None
        ws_summary.title = "Summary"
        self._write_summary(ws_summary, db, domain_id, org_id)

        # ── Sheet 2: Entities ──────────────────────────────────────────────────
        ws_entities = wb.create_sheet("Entities")
        self._write_entities(ws_entities, db, domain_id, org_id)

        # topic_clusters used to be written here by a bespoke sheet writer that
        # queried TopicAnalyzer directly with its own cap of 50, while PPTX used
        # 20 and HTML 15. It now goes through the shared collector below like
        # every other migrated section, so all four formats state the same
        # finding over the same rows.

        # ── Migrated sections: rendered from the shared section payload ───────
        # These render via the format-neutral collector + Excel renderer, so the
        # section is authored once and appears here without a bespoke writer.
        # (unify-report-format-coverage phase 3; entity_stats was the pilot.)
        # Each migrated section is one entry here — the strangler grows this map.
        # Requested sections are canonicalized so an alias (e.g. top_brands)
        # still resolves to its migrated collector.
        from backend import report_builder
        from backend.reporting.excel_renderer import render_excel
        requested = set(report_builder.canonical_sections(sections))
        migrated_collectors = {
            "entity_stats": report_builder.collect_entity_stats,
            "enrichment_coverage": report_builder.collect_enrichment_coverage,
            "top_secondary_labels": report_builder.collect_top_secondary_labels,
            "impact_projection": report_builder.collect_impact_projection,
            "institutional_benchmark": report_builder.collect_institutional_benchmark,
            "hidden_patterns": report_builder.collect_hidden_patterns,
            "decision_recommendations": report_builder.collect_decision_recommendations,
            "authority_control": report_builder.collect_authority_control,
            "collaboration_graph": report_builder.collect_collaboration_graph,
            "journal_portfolio": report_builder.collect_journal_portfolio,
            "topic_clusters": report_builder.collect_topic_clusters,
        }
        collected: list[tuple[str, object]] = []
        for section_id, collect in migrated_collectors.items():
            if section_id in requested:
                payload = collect(db, domain_id, org_id)
                sheet = render_excel(payload, wb)
                collected.append((sheet.title, payload))

        # ── Sheet 4: Harmonization Log ────────────────────────────────────────
        # Still a bespoke writer: its sheet carries columns the collector's payload
        # does not (row ids, executed-at, reverted) over up to 200 rows, and this
        # change is not the place to trade that detail away — 3.4 already showed
        # what migrating a section costs when the payload cap is lower than the
        # sheet's. But the parity requirement applies to what a format *renders*,
        # not to how it renders it, so the section's finding and disclosure are
        # collected for the Methodology sheet either way. Migrating the writer
        # itself remains open under report-format-parity.
        if "harmonization_log" in requested:
            payload = report_builder.collect_harmonization_log(db, domain_id, org_id)
            ws_harm = wb.create_sheet("Harmonization")
            self._write_harmonization(ws_harm, db, org_id, caveat=payload.method)
            collected.append((ws_harm.title, payload))

        # Built after the section sheets and moved to the front — the same shape
        # as the HTML executive summary, and for the same reason: it states what
        # every section found and cannot know that until each has collected.
        if collected:
            self._write_methodology(wb, collected)

        if manual_sections:
            ws_notes = wb.create_sheet("Analyst Notes")
            self._write_manual_sections(ws_notes, manual_sections)

        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ── Private sheet writers ──────────────────────────────────────────────────

    def _write_summary(self, ws, db: Session, domain_id: str, org_id: int | None) -> None:
        query = self._entities_query(db, domain_id, org_id)
        total = query.count()
        enriched = (
            query
            .filter(models.RawEntity.enrichment_status == "completed")
            .count()
        )
        pct = round(enriched / total * 100, 1) if total > 0 else 0.0

        from sqlalchemy import func
        avg_row = (
            query.with_entities(func.avg(models.RawEntity.enrichment_citation_count))
            .filter(models.RawEntity.enrichment_status == "completed")
            .scalar()
        )
        avg_cit = round(float(avg_row), 1) if avg_row is not None else 0.0

        headers = ["Metric", "Value"]
        _style_header_row(ws, headers)

        rows = [
            ("Active Domain",       domain_id),
            ("Total Entities",      total),
            ("Enriched Entities",   enriched),
            ("Enrichment %",        f"{pct}%"),
            ("Avg Citations",       avg_cit),
            ("Platform",            "UKIP — Universal Knowledge Intelligence Platform"),
        ]
        for row_idx, (metric, value) in enumerate(rows, start=2):
            ws.cell(row=row_idx, column=1, value=metric).font = _SUBHEADER_FONT
            ws.cell(row=row_idx, column=2, value=value)

        _autofit(ws)

    def _write_entities(self, ws, db: Session, domain_id: str, org_id: int | None) -> None:
        headers = [
            "ID", "Primary Label", "Secondary Label", "Canonical ID", "Entity Type",
            "Enrichment Status", "Citation Count", "Source",
        ]
        _style_header_row(ws, headers)
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        rows = (
            self._entities_query(db, domain_id, org_id)
            .order_by(models.RawEntity.id)
            .limit(self._ENTITY_CAP)
            .all()
        )
        for row_idx, e in enumerate(rows, start=2):
            ws.cell(row=row_idx, column=1, value=e.id)
            ws.cell(row=row_idx, column=2, value=e.primary_label)
            ws.cell(row=row_idx, column=3, value=e.secondary_label)
            ws.cell(row=row_idx, column=4, value=e.canonical_id)
            ws.cell(row=row_idx, column=5, value=e.entity_type)
            ws.cell(row=row_idx, column=6, value=e.enrichment_status)
            ws.cell(row=row_idx, column=7, value=e.enrichment_citation_count)
            ws.cell(row=row_idx, column=8, value=e.enrichment_source)

        _autofit(ws)

    def _write_harmonization(
        self, ws, db: Session, org_id: int | None, caveat: str | None = None
    ) -> None:
        headers = ["ID", "Step ID", "Step Name", "Records Updated", "Fields Modified", "Executed At", "Reverted"]
        header_row = 1
        if caveat:
            # Directly above the header, so a copied range carries it (5.3).
            cell = ws.cell(row=1, column=1, value=caveat)
            cell.font = _CAVEAT_FONT
            cell.alignment = _WRAP_TOP
            header_row = 2
        _style_header_row(ws, headers, row=header_row)

        rows = (
            self._harmonization_query(db, org_id)
            .order_by(models.HarmonizationLog.id.desc())
            .limit(200)
            .all()
        )
        for row_idx, h in enumerate(rows, start=header_row + 1):
            ws.cell(row=row_idx, column=1, value=h.id)
            ws.cell(row=row_idx, column=2, value=h.step_id)
            ws.cell(row=row_idx, column=3, value=h.step_name)
            ws.cell(row=row_idx, column=4, value=h.records_updated)
            ws.cell(row=row_idx, column=5, value=h.fields_modified)
            ws.cell(row=row_idx, column=6, value=str(h.executed_at) if h.executed_at else "")
            ws.cell(row=row_idx, column=7, value="Yes" if h.reverted else "No")

        _autofit(ws)

    def _write_methodology(self, wb, collected: list[tuple[str, object]]) -> None:
        """One row per rendered section: where to find it, what it found, and how.

        Keyed on the sheet name rather than an exhibit ordinal. A workbook renders
        a different set of sections than the document does — `agentic_trace` is
        declared unsupported here, and this exporter iterates its own collector
        map rather than the requested order — so a workbook numbering its own
        exhibits would agree with the PDF up to the first divergence and then be
        off by one for everything after, silently, from the same request. In a
        workbook the sheet tab is how a reader navigates anyway (design
        decision 7).

        Placed second, behind Summary: a disclosure a reader has to hunt for at
        the end of a twelve-tab workbook is one they will not read.
        """
        ws = wb.create_sheet("Methodology")
        _style_header_row(ws, ["Sheet", "Finding", "Source & caveat"])

        for row_idx, (sheet_name, payload) in enumerate(collected, start=2):
            ws.cell(row=row_idx, column=1, value=sheet_name).font = _SUBHEADER_FONT
            ws.cell(row=row_idx, column=2, value=payload.takeaway).alignment = _WRAP_TOP
            ws.cell(row=row_idx, column=3, value=payload.method).alignment = _WRAP_TOP

        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 52
        ws.column_dimensions["C"].width = 96
        wb.move_sheet(ws, offset=-(len(wb.sheetnames) - 2))

    def _write_manual_sections(self, ws, manual_sections: list[dict[str, str]]) -> None:
        headers = ["Section", "Analyst Text"]
        _style_header_row(ws, headers)
        for row_idx, section in enumerate(manual_sections, start=2):
            ws.cell(row=row_idx, column=1, value=(section.get("title") or "Analyst Note")[:120])
            cell = ws.cell(row=row_idx, column=2, value=section.get("content") or "")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 90
